"""Module for writing word data to Excel in the specified format."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def write_formatted_excel(
    words_df: pd.DataFrame,
    excel_path: Path,
    sheet_name: str,
) -> None:
    """
    Write the words dataframe to Excel in the specified alphabetical format.

    The format includes:
    - Header row: '', 'Nouns', '', 'Verbs', '', 'Adjectives', ''
    - First column: letter navigation (first letter uppercase, then second letters lowercase)
    - Word columns for each POS category followed by count columns
    - Each row contains words from different POS categories that share the same first and second letters
    - The second letter appears only once (on the first row for that letter combination)

    Args:
        words_df: DataFrame with columns 'word', 'POS', 'count'
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to create/overwrite
    """
    wb = load_workbook(excel_path)

    # Remove sheet if it exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Write header row
    ws.append(["", "Nouns", "", "Verbs", "", "Adjectives", ""])

    # Separate words by POS and create lookup dictionaries with lists
    nouns_df = words_df[words_df["POS"] == "noun"].sort_values(by="word")
    verbs_df = words_df[words_df["POS"] == "verb"].sort_values(by="word")
    adjs_df = words_df[words_df["POS"] == "adj"].sort_values(by="word")

    nouns_dict: dict[str, int] = dict(zip(nouns_df["word"], nouns_df["count"]))
    verbs_dict: dict[str, int] = dict(zip(verbs_df["word"], verbs_df["count"]))
    adjs_dict: dict[str, int] = dict(zip(adjs_df["word"], adjs_df["count"]))

    # Get all words sorted alphabetically
    all_words_sorted = sorted(
        set(nouns_dict.keys()) | set(verbs_dict.keys()) | set(adjs_dict.keys())
    )

    if not all_words_sorted:
        wb.save(excel_path)
        return

    # Group words by first letter, then by second letter, then by POS
    grouped: dict[str, dict[str, dict[str, list[str]]]] = {}

    for word in all_words_sorted:
        first_letter = word[0].upper()
        second_letter = word[1].lower() if len(word) > 1 else ""

        if first_letter not in grouped:
            grouped[first_letter] = {}
        if second_letter not in grouped[first_letter]:
            grouped[first_letter][second_letter] = {"noun": [], "verb": [], "adj": []}

        if word in nouns_dict:
            grouped[first_letter][second_letter]["noun"].append(word)
        if word in verbs_dict:
            grouped[first_letter][second_letter]["verb"].append(word)
        if word in adjs_dict:
            grouped[first_letter][second_letter]["adj"].append(word)

    # Write data row by row
    for first_letter in sorted(grouped.keys()):
        first_letter_row_written = False

        for second_letter in sorted(grouped[first_letter].keys()):
            pos_words = grouped[first_letter][second_letter]

            # Get the maximum number of words across all POS categories for this letter combo
            max_words = max(
                len(pos_words["noun"]), len(pos_words["verb"]), len(pos_words["adj"])
            )

            # Track if we've shown the second letter for this combination
            second_letter_shown = False

            # Write rows for each word index
            for word_idx in range(max_words):
                # Determine the first column value
                if not first_letter_row_written:
                    # Very first row for this first letter - show uppercase first letter
                    first_col = first_letter
                    first_letter_row_written = True
                elif not second_letter_shown:
                    # First row for this second letter combination - show lowercase second letter
                    first_col = second_letter
                    second_letter_shown = True
                else:
                    # Subsequent rows with same second letter - leave blank
                    first_col = ""

                # Get words and counts for this row
                noun_word = (
                    pos_words["noun"][word_idx]
                    if word_idx < len(pos_words["noun"])
                    else ""
                )
                noun_count = nouns_dict.get(noun_word, "") if noun_word else ""

                verb_word = (
                    pos_words["verb"][word_idx]
                    if word_idx < len(pos_words["verb"])
                    else ""
                )
                verb_count = verbs_dict.get(verb_word, "") if verb_word else ""

                adj_word = (
                    pos_words["adj"][word_idx]
                    if word_idx < len(pos_words["adj"])
                    else ""
                )
                adj_count = adjs_dict.get(adj_word, "") if adj_word else ""

                # Build and write row
                row_data = [
                    first_col,
                    noun_word,
                    noun_count,
                    verb_word,
                    verb_count,
                    adj_word,
                    adj_count,
                ]

                ws.append(row_data)

    wb.save(excel_path)
